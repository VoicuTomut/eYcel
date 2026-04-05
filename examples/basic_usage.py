#!/usr/bin/env python3
"""
eYcel Basic Usage Example

This example demonstrates:
1. Creating a sample Excel file
2. Encrypting it (anonymizing data while preserving formulas)
3. Decrypting it (restoring original values)
4. Verifying the round-trip worked correctly
"""
import os
import sys
import tempfile
from pathlib import Path

# Add parent directory to path so examples work without installation
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from eYcel import encrypt_excel, decrypt_excel
from eYcel.formula_handler import verify_formulas_preserved


def create_sample_excel(filepath: str) -> None:
    """Create a sample Excel file with data and formulas."""
    from openpyxl import Workbook
    from datetime import datetime, timedelta
    import random

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Data"

    # Headers
    headers = ["Date", "Product", "Region", "Quantity", "Price", "Revenue"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Sample data
    products = ["Widget", "Gadget", "Tool", "Device"]
    regions = ["North", "South", "East", "West"]
    base_date = datetime(2024, 1, 1)

    for row in range(2, 12):  # 10 rows of data
        # Date
        date_val = base_date + timedelta(days=row - 2)
        ws.cell(row=row, column=1, value=date_val)

        # Product (categorical)
        ws.cell(row=row, column=2, value=random.choice(products))

        # Region (categorical)
        ws.cell(row=row, column=3, value=random.choice(regions))

        # Quantity (integer)
        ws.cell(row=row, column=4, value=random.randint(10, 100))

        # Price (float)
        ws.cell(row=row, column=5, value=round(random.uniform(10.0, 100.0), 2))

        # Revenue (formula = Quantity * Price)
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")

    # Add a totals row with formulas
    ws.cell(row=12, column=1, value="TOTAL")
    ws.cell(row=12, column=4, value="=SUM(D2:D11)")
    ws.cell(row=12, column=5, value="=SUM(E2:E11)")
    ws.cell(row=12, column=6, value="=SUM(F2:F11)")

    wb.save(filepath)
    print(f"✓ Created sample file: {filepath}")


def main():
    """Run the encrypt/decrypt demonstration."""
    print("=" * 60)
    print("eYcel Basic Usage Example")
    print("=" * 60)

    # Use a temporary directory for this example
    with tempfile.TemporaryDirectory() as tmpdir:
        # Define file paths
        original_file = os.path.join(tmpdir, "sample_data.xlsx")
        encrypted_file = os.path.join(tmpdir, "sample_encrypted.xlsx")
        rules_file = os.path.join(tmpdir, "sample_rules.yaml")
        restored_file = os.path.join(tmpdir, "sample_restored.xlsx")

        # Step 1: Create sample Excel file
        print("\n1. Creating sample Excel file...")
        create_sample_excel(original_file)

        # Step 2: Encrypt the file
        print("\n2. Encrypting (anonymizing) the file...")
        try:
            encrypt_excel(original_file, encrypted_file)
            print(f"✓ Encrypted file created: {encrypted_file}")
            print(f"✓ Rules file created: {rules_file}")
        except Exception as e:
            print(f"✗ Encryption failed: {e}")
            return

        # Step 3: Decrypt the file
        print("\n3. Decrypting (restoring) the file...")
        try:
            decrypt_excel(encrypted_file, rules_file, restored_file)
            print(f"✓ Restored file created: {restored_file}")
        except Exception as e:
            print(f"✗ Decryption failed: {e}")
            return

        # Step 4: Verify formulas preserved
        print("\n4. Verifying formula preservation...")
        try:
            formulas_ok = verify_formulas_preserved(original_file, restored_file)
            if formulas_ok:
                print("✓ All formulas preserved correctly!")
            else:
                print("✗ Formula mismatch detected!")
        except Exception as e:
            print(f"✗ Formula verification failed: {e}")

        # Summary
        print("\n" + "=" * 60)
        print("Round-trip Summary:")
        print("=" * 60)
        print(f"Original:   {original_file}")
        print(f"Encrypted:  {encrypted_file}")
        print(f"Rules:      {rules_file}")
        print(f"Restored:   {restored_file}")
        print("\n✓ Example completed successfully!")

        # File sizes
        orig_size = os.path.getsize(original_file)
        enc_size = os.path.getsize(encrypted_file)
        print(f"\nFile sizes:")
        print(f"  Original:  {orig_size:,} bytes")
        print(f"  Encrypted: {enc_size:,} bytes")


if __name__ == "__main__":
    main()
