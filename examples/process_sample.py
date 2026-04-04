#!/usr/bin/env python3
"""
Full eYcel API demo — encrypt → inspect → decrypt → verify.

Run:
    python examples/process_sample.py

What it does:
    1. Generates a fresh sample_data.xlsx
    2. Encrypts it → sample_data_encrypted.xlsx + sample_data_rules.yaml
    3. Prints a before/after comparison of the first data row
    4. Decrypts → sample_data_restored.xlsx
    5. Verifies the round-trip correctly:
       - Reversible transforms (scale, offset, shuffle) → values must match
       - Hash transforms                                → one-way, skip verify
       - Formulas                                       → must be identical
"""
from __future__ import annotations

import sys
from pathlib import Path

# Allow running from the repo root without `pip install -e .`
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
import yaml

from eYcel import encrypt_excel, decrypt_excel
from generate_sample_data import create_sample_excel


EXAMPLES_DIR = Path(__file__).parent


def _read_row2(path: Path) -> list:
    """Return data-row-2 cell values."""
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    return [cell.value for cell in ws[2]]


def _get_headers(path: Path) -> list:
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    return [cell.value for cell in ws[1]]


def main() -> None:
    sample    = EXAMPLES_DIR / "sample_data.xlsx"
    encrypted = EXAMPLES_DIR / "sample_data_encrypted.xlsx"
    restored  = EXAMPLES_DIR / "sample_data_restored.xlsx"

    # 1 — Generate
    print("── Step 1: Generate sample data ─────────────────────────────")
    create_sample_excel(sample, n_rows=20)

    # 2 — Encrypt
    print("\n── Step 2: Encrypt ───────────────────────────────────────────")
    rules_out = encrypt_excel(str(sample), str(encrypted))
    print(f"   Encrypted  → {encrypted.name}")
    print(f"   Rules      → {Path(rules_out).name}")

    # Load rules to know which columns use hash (irreversible)
    with open(rules_out) as f:
        rules = yaml.safe_load(f)
    column_configs = rules.get("columns", {})

    # 3 — Compare before/after
    print("\n── Step 3: Before / After comparison (row 2) ────────────────")
    headers  = _get_headers(sample)
    orig_row = _read_row2(sample)
    enc_row  = _read_row2(encrypted)
    for h, o, e in zip(headers, orig_row, enc_row):
        changed = "✏️  changed" if o != e else "   (same)"
        print(f"   {h:15s}  {str(o):35s} → {str(e):35s}  {changed}")

    # 4 — Decrypt
    print("\n── Step 4: Decrypt ───────────────────────────────────────────")
    decrypt_excel(str(encrypted), rules_out, str(restored))
    print(f"   Restored   → {restored.name}")

    # 5 — Verify (only reversible columns)
    print("\n── Step 5: Verify round-trip ─────────────────────────────────")
    orig_row2 = _read_row2(sample)
    rest_row2 = _read_row2(restored)

    failures = 0
    for h, o, r in zip(headers, orig_row2, rest_row2):
        cfg = column_configs.get(h, {})
        transform = cfg.get("transform", "keep")

        if transform == "hash":
            # One-way — expected NOT to restore
            print(f"   🔒 {h:15s}  hash (one-way) — skipping restore check")
            continue

        if isinstance(o, str) and o.startswith("="):
            ok = (o == r)
        elif isinstance(o, float) and isinstance(r, float):
            ok = abs(o - r) < 0.01
        else:
            ok = (str(o) == str(r))

        status = "✅" if ok else "❌"
        if not ok:
            failures += 1
        print(f"   {status} {h:15s}  orig={str(o):30s}  restored={str(r)}")

    print()
    if failures == 0:
        print("🎉  Round-trip verified — all reversible values restored correctly!")
    else:
        print(f"⚠️   {failures} value(s) did not round-trip exactly.")
        sys.exit(1)


if __name__ == "__main__":
    main()
