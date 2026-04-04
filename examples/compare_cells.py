import openpyxl
import sys

def compare_workbooks(path1, path2):
    wb1 = openpyxl.load_workbook(path1, data_only=False)
    wb2 = openpyxl.load_workbook(path2, data_only=False)
    
    if wb1.sheetnames != wb2.sheetnames:
        print(f"Sheet names differ: {wb1.sheetnames} vs {wb2.sheetnames}")
        return False
    
    for sheet_name in wb1.sheetnames:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        
        if ws1.max_row != ws2.max_row:
            print(f"Row count mismatch in {sheet_name}: {ws1.max_row} vs {ws2.max_row}")
            return False
        if ws1.max_column != ws2.max_column:
            print(f"Column count mismatch in {sheet_name}: {ws1.max_column} vs {ws2.max_column}")
            return False
        
        for row in range(1, ws1.max_row + 1):
            for col in range(1, ws1.max_column + 1):
                cell1 = ws1.cell(row=row, column=col)
                cell2 = ws2.cell(row=row, column=col)
                
                # Compare data type
                if cell1.data_type != cell2.data_type:
                    print(f"Data type mismatch at {sheet_name}!{cell1.coordinate}: {cell1.data_type} vs {cell2.data_type}")
                    return False
                
                # Compare value
                if cell1.value != cell2.value:
                    print(f"Value mismatch at {sheet_name}!{cell1.coordinate}: {cell1.value} vs {cell2.value}")
                    return False
                
                # For formula cells, compare formula string
                if cell1.data_type == 'f' and cell1.value != cell2.value:
                    print(f"Formula mismatch at {sheet_name}!{cell1.coordinate}: {cell1.value} vs {cell2.value}")
                    return False
    
    print("All cells match.")
    return True

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python compare_cells.py file1.xlsx file2.xlsx")
        sys.exit(1)
    ok = compare_workbooks(sys.argv[1], sys.argv[2])
    sys.exit(0 if ok else 1)