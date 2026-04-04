"""
eYcel Streamlit GUI Application.

A web interface for encrypting and decrypting Excel files while
preserving formulas and structure.
"""
import os
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import streamlit as st
import openpyxl
import yaml

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from eYcel.column_analyzer import analyze_workbook_columns
from eYcel.formula_handler import extract_formulas, get_formula_summary
from eYcel.encrypt import encrypt_excel, auto_detect_transform
from eYcel.decrypt import decrypt_excel, load_and_validate_rules
from eYcel.yaml_handler import validate_rules


# ---------------------------------------------------------------------------
# Constants & Configuration
# ---------------------------------------------------------------------------

APP_TITLE = "🔐 eYcel"
APP_DESCRIPTION = "Excel encryption & decryption with formula preservation"
MAX_PREVIEW_ROWS = 5

TRANSFORM_ICONS = {
    "hash": "🔒",
    "offset": "📅",
    "scale": "📊",
    "shuffle": "🔀",
    "keep": "✋",
    "anonymize": "👤",
}

TRANSFORM_DESCRIPTIONS = {
    "hash": "One-way hash (irreversible)",
    "offset": "Date/number offset (reversible)",
    "scale": "Numeric scaling (reversible)",
    "shuffle": "Category shuffle (reversible)",
    "keep": "No change",
    "anonymize": "Pattern-based anonymization",
}


# ---------------------------------------------------------------------------
# Session State Management
# ---------------------------------------------------------------------------

def init_session_state() -> None:
    """Initialize Streamlit session state variables."""
    defaults = {
        "page": "Encrypt",
        "temp_files": [],
        "encrypted_file": None,
        "rules_file": None,
        "decrypted_file": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def cleanup_temp_files() -> None:
    """Remove all tracked temporary files."""
    for filepath in st.session_state.temp_files:
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception:
            pass
    st.session_state.temp_files = []


def register_temp_file(filepath: str) -> str:
    """Register a temp file for cleanup and return the path."""
    st.session_state.temp_files.append(filepath)
    return filepath


# ---------------------------------------------------------------------------
# File Handling Utilities
# ---------------------------------------------------------------------------

def save_uploaded_file(uploaded_file) -> Optional[str]:
    """Save an uploaded file to a temporary location.
    
    Args:
        uploaded_file: Streamlit UploadedFile object.
        
    Returns:
        Path to the saved temp file, or None on failure.
    """
    if uploaded_file is None:
        return None
    try:
        suffix = Path(uploaded_file.name).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded_file.getvalue())
            return register_temp_file(tmp.name)
    except Exception as e:
        st.error(f"Failed to save uploaded file: {e}")
        return None


def get_preview_data(filepath: str, max_rows: int = MAX_PREVIEW_ROWS) -> Tuple[List[str], List[List[Any]]]:
    """Extract preview data from an Excel file.
    
    Args:
        filepath: Path to Excel file.
        max_rows: Maximum rows to include.
        
    Returns:
        Tuple of (headers, rows) for preview display.
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.active
    
    headers = []
    rows = []
    
    for row_idx, row in enumerate(ws.iter_rows(max_row=min(max_rows + 1, ws.max_row)), 1):
        values = [cell.value for cell in row]
        if row_idx == 1:
            headers = values
        else:
            rows.append(values)
    
    return headers, rows


def count_formulas_in_file(filepath: str) -> int:
    """Count total formulas across all sheets in a workbook.
    
    Args:
        filepath: Path to Excel file.
        
    Returns:
        Total formula count.
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total += get_formula_summary(ws)["formula_count"]
    return total


# ---------------------------------------------------------------------------
# Encryption Helpers
# ---------------------------------------------------------------------------

def show_preview_section(temp_path: str) -> bool:
    """Show file preview section. Returns success status."""
    st.subheader("📋 File Preview")
    try:
        headers, rows = get_preview_data(temp_path)
        preview_data = {h: [r[i] if i < len(r) else None for r in rows] 
                       for i, h in enumerate(headers)}
        st.dataframe(preview_data, use_container_width=True)
        return True
    except Exception as e:
        st.error(f"Failed to preview file: {e}")
        return False


def analyze_and_show_columns(temp_path: str) -> Optional[Dict]:
    """Analyze columns and show results. Returns analysis dict or None."""
    st.subheader("📊 Column Analysis")
    try:
        wb = openpyxl.load_workbook(temp_path, data_only=False)
        analysis = analyze_workbook_columns(wb, wb.sheetnames[0])
        
        analysis_data = []
        for col_letter, meta in analysis.items():
            transform = auto_detect_transform(meta)
            icon = TRANSFORM_ICONS.get(transform, "❓")
            desc = TRANSFORM_DESCRIPTIONS.get(transform, transform)
            
            analysis_data.append({
                "Column": col_letter,
                "Header": meta.get("header", "—"),
                "Type": meta.get("dominant_type", "unknown"),
                "Suggested": f"{icon} {transform}",
                "Description": desc,
                "Formulas": meta.get("formula_count", 0),
                "Data Cells": meta.get("data_count", 0),
            })
        
        st.dataframe(analysis_data, use_container_width=True, hide_index=True)
        
        formula_count = sum(m.get("formula_count", 0) for m in analysis.values())
        data_count = sum(m.get("data_count", 0) for m in analysis.values())
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Columns", len(analysis))
        col2.metric("Formula Cells", formula_count)
        col3.metric("Data Cells", data_count)
        
        return analysis
    except Exception as e:
        st.error(f"Failed to analyze columns: {e}")
        return None


def perform_encryption(temp_path: str, uploaded_name: str) -> Optional[Tuple[str, str]]:
    """Execute encryption and return (encrypted_path, rules_path)."""
    output_dir = tempfile.gettempdir()
    base_name = Path(uploaded_name).stem
    encrypted_path = os.path.join(output_dir, f"{base_name}_encrypted.xlsx")
    
    rules_path = encrypt_excel(temp_path, encrypted_path)
    return encrypted_path, rules_path


def show_encrypt_downloads(encrypted_path: str, rules_path: str, base_name: str) -> None:
    """Show download buttons for encrypted file and rules."""
    col1, col2 = st.columns(2)
    
    with open(encrypted_path, "rb") as f:
        col1.download_button(
            label="📥 Download Encrypted File",
            data=f,
            file_name=f"{base_name}_encrypted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    
    with open(rules_path, "rb") as f:
        col2.download_button(
            label="📥 Download Rules File",
            data=f,
            file_name=f"{base_name}_rules.yaml",
            mime="application/x-yaml",
            use_container_width=True,
        )


def show_encrypt_summary(analysis: Dict, encrypted_path: str) -> None:
    """Show encryption summary metrics."""
    st.subheader("📈 Encryption Summary")
    encrypted_formulas = count_formulas_in_file(encrypted_path)
    
    sum_col1, sum_col2, sum_col3 = st.columns(3)
    sum_col1.metric("Columns Processed", len(analysis))
    sum_col2.metric("Formulas Preserved", encrypted_formulas)
    sum_col3.metric("Rules Generated", "✓")
    
    st.info("💾 **Save both files!** The rules file is required for decryption.")


# ---------------------------------------------------------------------------
# Decryption Helpers
# ---------------------------------------------------------------------------

def validate_rules_content(rules_file) -> Optional[Dict]:
    """Validate uploaded rules file content. Returns rules dict or None."""
    try:
        rules_content = yaml.safe_load(rules_file.getvalue())
        is_valid, errors = validate_rules(rules_content)
        
        if not is_valid:
            st.error("❌ Invalid rules file:")
            for err in errors:
                st.write(f"  • {err}")
            return None
        
        st.success("✅ Rules file validated")
        return rules_content
    except Exception as e:
        st.error(f"Failed to validate rules: {e}")
        return None


def show_rules_metadata(rules_content: Dict) -> None:
    """Display rules metadata in columns."""
    meta = rules_content.get("metadata", {})
    col_meta = st.columns(3)
    col_meta[0].metric("Source File", meta.get("original_filename", "Unknown"))
    col_meta[1].metric("Created", meta.get("timestamp", "Unknown")[:10] if meta.get("timestamp") else "Unknown")
    col_meta[2].metric("Columns", len(rules_content.get("columns", {})))


def perform_decryption(
    encrypted_path: str, 
    rules_path: str, 
    base_name: str
) -> Optional[str]:
    """Execute decryption and return decrypted file path."""
    output_dir = tempfile.gettempdir()
    decrypted_path = os.path.join(output_dir, f"{base_name}_decrypted.xlsx")
    register_temp_file(decrypted_path)
    
    decrypt_excel(encrypted_path, rules_path, decrypted_path)
    return decrypted_path


def show_decrypt_download(decrypted_path: str, base_name: str) -> None:
    """Show download button for decrypted file."""
    with open(decrypted_path, "rb") as f:
        st.download_button(
            label="📥 Download Decrypted File",
            data=f,
            file_name=f"{base_name}_decrypted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def show_verification_results(encrypted_path: str, decrypted_path: str) -> None:
    """Show verification metrics comparing encrypted and decrypted files."""
    st.subheader("✅ Verification Results")
    
    wb_orig = openpyxl.load_workbook(encrypted_path, data_only=False)
    wb_dec = openpyxl.load_workbook(decrypted_path, data_only=False)
    
    ver_col1, ver_col2, ver_col3 = st.columns(3)
    
    orig_formulas = count_formulas_in_file(encrypted_path)
    dec_formulas = count_formulas_in_file(decrypted_path)
    formulas_match = orig_formulas == dec_formulas
    ver_col1.metric("Formulas", "✓ Match" if formulas_match else "✗ Mismatch",
                   f"{dec_formulas} formulas")
    
    orig_rows = wb_orig.active.max_row
    dec_rows = wb_dec.active.max_row
    rows_match = orig_rows == dec_rows
    ver_col2.metric("Row Count", "✓ Match" if rows_match else "✗ Mismatch",
                   f"{dec_rows} rows")
    
    orig_cols = wb_orig.active.max_column
    dec_cols = wb_dec.active.max_column
    cols_match = orig_cols == dec_cols
    ver_col3.metric("Column Count", "✓ Match" if cols_match else "✗ Mismatch",
                   f"{dec_cols} columns")
    
    if formulas_match and rows_match and cols_match:
        st.balloons()


# ---------------------------------------------------------------------------
# UI Page Renderers
# ---------------------------------------------------------------------------

def render_sidebar() -> str:
    """Render the sidebar navigation and return selected page.
    
    Returns:
        Selected page name.
    """
    with st.sidebar:
        st.title(APP_TITLE)
        st.caption(APP_DESCRIPTION)
        st.divider()
        
        page = st.radio(
            "Navigate",
            ["🔐 Encrypt", "🔓 Decrypt", "📋 Validate Rules"],
            format_func=lambda x: x,
        )
        
        st.divider()
        
        with st.expander("About eYcel"):
            st.markdown("""
            **eYcel** anonymizes Excel data while preserving formulas.
            
            **Features:**
            - 🔒 Hash strings (one-way)
            - 📅 Offset dates/numbers (reversible)
            - 📊 Scale numeric values (reversible)
            - 🔀 Shuffle categories (reversible)
            - ✋ Keep values unchanged
            
            All formulas are extracted, preserved, and restored exactly.
            """)
        
        st.divider()
        st.caption("v1.0.0 | Phase 5")
        
    return page.replace("🔐 ", "").replace("🔓 ", "").replace("📋 ", "")


def render_encrypt_page() -> None:
    """Render the Encrypt page with file upload, preview, and encryption."""
    st.header("🔐 Encrypt Excel File")
    st.markdown("Upload an Excel file to encrypt sensitive data while preserving formulas.")
    
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=["xlsx", "xls", "xlsm"],
        key="encrypt_uploader",
    )
    
    if uploaded_file is None:
        st.info("📤 Upload an Excel file to get started.")
        return
    
    temp_path = save_uploaded_file(uploaded_file)
    if temp_path is None:
        return
    
    if not show_preview_section(temp_path):
        return
    
    analysis = analyze_and_show_columns(temp_path)
    if analysis is None:
        return
    
    st.divider()
    
    if st.button("🔐 Encrypt File", type="primary", use_container_width=True):
        progress_bar = st.progress(0, text="Starting encryption...")
        
        try:
            progress_bar.progress(50, text="Processing...")
            encrypted_path, rules_path = perform_encryption(temp_path, uploaded_file.name)
            progress_bar.progress(100, text="Complete!")
            
            st.session_state.encrypted_file = encrypted_path
            st.session_state.rules_file = rules_path
            
            st.success("✅ Encryption complete!")
            
            base_name = Path(uploaded_file.name).stem
            show_encrypt_downloads(encrypted_path, rules_path, base_name)
            show_encrypt_summary(analysis, encrypted_path)
            
        except Exception as e:
            st.error(f"Encryption failed: {e}")
            progress_bar.empty()


def render_decrypt_page() -> None:
    """Render the Decrypt page with file upload and decryption."""
    st.header("🔓 Decrypt Excel File")
    st.markdown("Restore an encrypted Excel file using its rules file.")
    
    col1, col2 = st.columns(2)
    
    with col1:
        encrypted_file = st.file_uploader(
            "Encrypted Excel file",
            type=["xlsx", "xls", "xlsm"],
            key="decrypt_excel_uploader",
        )
    
    with col2:
        rules_file = st.file_uploader(
            "Rules YAML file",
            type=["yaml", "yml"],
            key="decrypt_rules_uploader",
        )
    
    if encrypted_file is None or rules_file is None:
        st.info("📤 Upload both the encrypted Excel file and the rules YAML file.")
        return
    
    rules_content = validate_rules_content(rules_file)
    if rules_content is None:
        return
    
    show_rules_metadata(rules_content)
    
    st.divider()
    
    if st.button("🔓 Decrypt File", type="primary", use_container_width=True):
        progress_bar = st.progress(0, text="Starting decryption...")
        
        try:
            encrypted_path = save_uploaded_file(encrypted_file)
            rules_path = save_uploaded_file(rules_file)
            
            if encrypted_path is None or rules_path is None:
                return
            
            progress_bar.progress(50, text="Processing...")
            base_name = Path(encrypted_file.name).stem.replace("_encrypted", "")
            decrypted_path = perform_decryption(encrypted_path, rules_path, base_name)
            progress_bar.progress(100, text="Complete!")
            
            st.session_state.decrypted_file = decrypted_path
            
            st.success("✅ Decryption complete!")
            show_decrypt_download(decrypted_path, base_name)
            show_verification_results(encrypted_path, decrypted_path)
            
        except Exception as e:
            st.error(f"Decryption failed: {e}")
            progress_bar.empty()


def render_validate_page() -> None:
    """Render the Validate Rules page."""
    st.header("📋 Validate Rules File")
    st.markdown("Inspect and validate a rules YAML file.")
    
    uploaded_file = st.file_uploader(
        "Choose rules YAML file",
        type=["yaml", "yml"],
        key="validate_uploader",
    )
    
    if uploaded_file is None:
        st.info("📤 Upload a rules YAML file to validate.")
        return
    
    try:
        rules_content = yaml.safe_load(uploaded_file.getvalue())
        is_valid, errors = validate_rules(rules_content)
        
        if is_valid:
            st.success("✅ Rules file is valid")
        else:
            st.error("❌ Rules file is invalid")
            for err in errors:
                st.write(f"  • {err}")
        
        st.subheader("📄 Metadata")
        meta = rules_content.get("metadata", {})
        meta_data = [{"Key": k, "Value": str(v)} for k, v in meta.items()]
        
        if meta_data:
            st.dataframe(meta_data, use_container_width=True, hide_index=True)
        else:
            st.info("No metadata found")
        
        st.subheader("🔧 Column Transformations")
        columns = rules_content.get("columns", {})
        
        if not columns:
            st.info("No column transformations defined")
        else:
            col_data = []
            for col_name, config in columns.items():
                transform = config.get("transform", "unknown")
                icon = TRANSFORM_ICONS.get(transform, "❓")
                
                params = []
                for k, v in config.items():
                    if k != "transform":
                        if isinstance(v, dict):
                            params.append(f"{k}={len(v)} entries")
                        else:
                            params.append(f"{k}={v}")
                
                col_data.append({
                    "Column": col_name,
                    "Transform": f"{icon} {transform}",
                    "Parameters": ", ".join(params) if params else "—",
                })
            
            st.dataframe(col_data, use_container_width=True, hide_index=True)
            st.metric("Total Columns", len(columns))
        
        with st.expander("📝 View Raw YAML"):
            st.code(uploaded_file.getvalue().decode("utf-8"), language="yaml")
            
    except yaml.YAMLError as e:
        st.error(f"Invalid YAML syntax: {e}")
    except Exception as e:
        st.error(f"Failed to process file: {e}")


def main() -> None:
    """Main application entry point."""
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="🔐",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    
    init_session_state()
    page = render_sidebar()
    st.session_state.page = page
    
    if page == "Encrypt":
        render_encrypt_page()
    elif page == "Decrypt":
        render_decrypt_page()
    elif page == "Validate Rules":
        render_validate_page()


if __name__ == "__main__":
    main()
