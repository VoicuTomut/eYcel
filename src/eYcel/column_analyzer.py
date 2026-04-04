"""
Column analyzer module for detecting data types and properties.
Analyzes openpyxl worksheets to classify each column's data type
and identify formula cells.
"""
from typing import Any, Dict, List, Optional, Union
from datetime import datetime, date


# ---------------------------------------------------------------------------
# Cell-level helpers
# ---------------------------------------------------------------------------

def detect_cell_type(cell_value: Any) -> str:
    """
    Detect the type of a single cell value.

    Args:
        cell_value: The value to analyse (raw Python value from openpyxl).

    Returns:
        One of: 'date', 'int', 'float', 'percentage', 'string',
                'categorical', 'boolean', or 'empty'.
    """
    if cell_value is None or cell_value == "":
        return "empty"
    if isinstance(cell_value, bool):
        return "boolean"
    if isinstance(cell_value, (datetime, date)):
        return "date"
    if isinstance(cell_value, int):
        return "int"
    if isinstance(cell_value, float):
        return "float"
    if isinstance(cell_value, str):
        return "string"
    return "string"


def is_formula_cell(cell) -> bool:
    """
    Check whether an openpyxl cell contains a formula.

    Args:
        cell: openpyxl Cell object.

    Returns:
        True if the cell value starts with '='.
    """
    return isinstance(cell.value, str) and cell.value.startswith("=")


# ---------------------------------------------------------------------------
# Column-level analysis
# ---------------------------------------------------------------------------

def detect_categorical(values: List[Any], threshold: float = 0.35) -> bool:
    """
    Decide if a column is categorical based on the unique-value ratio.

    A column is considered categorical when the ratio of distinct non-null
    values to total non-null values is below *threshold*.

    Args:
        values:    List of cell values (nulls allowed).
        threshold: Maximum unique/total ratio to classify as categorical.

    Returns:
        True if the column looks categorical.
    """
    non_null = [v for v in values if v is not None and v != ""]
    if len(non_null) == 0:
        return False
    unique_ratio = len(set(str(v) for v in non_null)) / len(non_null)
    return unique_ratio <= threshold


def get_column_stats(values: List[Union[int, float]]) -> Dict[str, float]:
    """
    Calculate basic statistics for a numeric column.

    Args:
        values: List of numeric values (non-null).

    Returns:
        Dict with keys: min, max, avg, unique_count.
    """
    if not values:
        return {"min": 0.0, "max": 0.0, "avg": 0.0, "unique_count": 0}
    numeric = [float(v) for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)]
    if not numeric:
        return {"min": 0.0, "max": 0.0, "avg": 0.0, "unique_count": 0}
    return {
        "min": min(numeric),
        "max": max(numeric),
        "avg": sum(numeric) / len(numeric),
        "unique_count": float(len(set(numeric))),
    }


def analyze_column(worksheet, column_letter: str) -> Dict[str, Any]:
    """
    Analyse all cells in a column and return rich metadata.

    Args:
        worksheet:     openpyxl Worksheet object (must be loaded with
                       data_only=False so formula strings are accessible).
        column_letter: Excel column letter, e.g. 'A'.

    Returns:
        Dictionary with keys:
          - letter        : str  — column letter
          - header        : Any  — value of row-1 cell
          - dominant_type : str  — most common non-empty cell type
          - is_categorical: bool — True if low cardinality string column
          - formula_count : int  — number of formula cells
          - data_count    : int  — number of non-formula, non-empty cells
          - stats         : dict — numeric stats (empty dict for non-numeric)
          - sample_values : list — up to 5 non-null, non-formula values
    """
    from openpyxl.utils import column_index_from_string

    col_idx = column_index_from_string(column_letter)
    max_row = worksheet.max_row or 1

    header = worksheet.cell(row=1, column=col_idx).value
    formula_count = 0
    type_counts: Dict[str, int] = {}
    raw_values: List[Any] = []

    for row in range(2, max_row + 1):
        cell = worksheet.cell(row=row, column=col_idx)
        if is_formula_cell(cell):
            formula_count += 1
            continue
        val = cell.value
        if val is None or val == "":
            continue
        ctype = detect_cell_type(val)
        type_counts[ctype] = type_counts.get(ctype, 0) + 1
        raw_values.append(val)

    dominant_type = max(type_counts, key=type_counts.get) if type_counts else "empty"
    is_cat = False
    if dominant_type == "string":
        is_cat = detect_categorical(raw_values)
        if is_cat:
            dominant_type = "categorical"

    stats: Dict[str, float] = {}
    if dominant_type in ("int", "float"):
        stats = get_column_stats(raw_values)

    return {
        "letter": column_letter,
        "header": header,
        "dominant_type": dominant_type,
        "is_categorical": is_cat,
        "formula_count": formula_count,
        "data_count": len(raw_values),
        "stats": stats,
        "sample_values": raw_values[:5],
    }


def analyze_workbook_columns(
    workbook, sheet_name: Optional[str] = None
) -> Dict[str, Dict[str, Any]]:
    """
    Analyse every column in a worksheet and return a mapping of
    column_letter → metadata dict.

    Args:
        workbook:   openpyxl Workbook (opened with data_only=False).
        sheet_name: Name of the sheet to analyse; defaults to the active sheet.

    Returns:
        Dict mapping column letters to their metadata dicts.
    """
    from openpyxl.utils import get_column_letter

    ws = workbook[sheet_name] if sheet_name else workbook.active
    max_col = ws.max_column or 1
    result: Dict[str, Dict[str, Any]] = {}

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        result[letter] = analyze_column(ws, letter)

    return result
