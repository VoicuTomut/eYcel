"""
Excel encryption pipeline.

Workflow
--------
1. Open the workbook (data_only=False so formulas are readable).
2. First pass: collect ALL unique text values to build global substitution map.
3. For each sheet, extract formulas, transform ALL cells, reinject formulas
   with text literals updated.
4. Save the encrypted workbook and rules.yaml.

Default behavior:
- ALL text (headers, titles, data) → consistent substitution (same word = same fake word)
- Text inside formulas → also substituted
- Numbers → kept unchanged (AI needs real values for formulas)
- Dates → kept unchanged
- Formulas → structure preserved, text literals inside updated
"""

import random
import string
from pathlib import Path
from typing import Any, Dict, Optional, Tuple
from datetime import datetime, timezone

import openpyxl
from openpyxl.utils import get_column_letter


from .formula_handler import extract_formulas, clear_formula_cells
from .transformations import (
    transform_hash,
    transform_offset_date,
    transform_offset_number,
    transform_scale,
    transform_shuffle,
    transform_keep,
    transform_anonymize,
    build_global_text_map,
    substitute_text_in_formula,
)
from .yaml_handler import generate_rules, save_rules


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _random_salt(length: int = 16) -> str:
    """Generate a random alphanumeric salt string."""
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


def _random_factor() -> float:
    """Return a random scale factor in [0.1, 0.9) ∪ (1.1, 2.0]."""
    factor = random.uniform(0.1, 2.0)
    while 0.9 <= factor <= 1.1:
        factor = random.uniform(0.1, 2.0)
    return round(factor, 6)


def _random_offset() -> float:
    """Return a random numeric offset in [-1000, -1) ∪ (1, 1000]."""
    return round(random.uniform(100, 1000) * random.choice([-1, 1]), 4)


def _random_day_offset() -> int:
    """Return a random day offset between ±365 days."""
    return random.randint(-365, 365)


def auto_detect_transform(
    col_values: list, scramble_numbers: bool = False, scramble_dates: bool = False,
) -> str:
    """
    Recommend a transform type based on cell values.

    New defaults:
    - string → "substitute" (consistent readable replacement)
    - int/float → "keep" (unless scramble_numbers=True → "scale")
    - date → "keep" (unless scramble_dates=True → "offset")
    """
    from .column_analyzer import detect_cell_type

    type_counts: Dict[str, int] = {}
    for v in col_values:
        if v is None or v == "":
            continue
        if isinstance(v, str) and v.startswith("="):
            continue
        ctype = detect_cell_type(v)
        type_counts[ctype] = type_counts.get(ctype, 0) + 1

    if not type_counts:
        return "keep"

    dominant = max(type_counts, key=type_counts.get)

    if dominant == "date":
        return "offset" if scramble_dates else "keep"
    if dominant in ("int", "float"):
        return "scale" if scramble_numbers else "keep"
    if dominant in ("string", "categorical"):
        return "substitute"
    return "keep"


def generate_output_paths(input_path: str) -> Tuple[str, str]:
    """Generate default output paths based on the input file name."""
    p = Path(input_path)
    stem = p.stem
    parent = p.parent
    encrypted = str(parent / f"{stem}_encrypted.xlsx")
    rules = str(parent / f"{stem}_rules.yaml")
    return encrypted, rules


# ---------------------------------------------------------------------------
# Column-level transform dispatcher
# ---------------------------------------------------------------------------

def _transform_cell(value: Any, config: Dict[str, Any], global_text_map: Dict[str, str]) -> Any:
    """Apply the correct forward transform to a single cell value."""
    # Text cells are ALWAYS substituted via the global map, regardless of column transform.
    # This ensures headers, titles, and text in numeric columns all get hidden.
    if isinstance(value, str) and not value.startswith("="):
        if value in global_text_map:
            return global_text_map[value]
        return value

    t = config.get("transform", "keep")
    if t == "hash":
        return transform_hash(value, config["salt"])
    if t == "offset":
        if hasattr(value, "year"):          # date / datetime
            return transform_offset_date(value, config["offset_days"])
        return transform_offset_number(value, config["offset"])
    if t == "scale":
        return transform_scale(value, config["factor"])
    if t == "shuffle":
        return transform_shuffle(value, config["mapping"])
    if t == "anonymize":
        return transform_anonymize(value, config.get("col_type", "string"))
    return transform_keep(value)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _load_workbook_any_format(path: Path):
    """Load .xlsx, .xls, or .csv into an openpyxl Workbook."""
    suffix = path.suffix.lower()

    if suffix == ".xls":
        # Convert .xls → openpyxl Workbook via xlrd
        import xlrd
        xls = xlrd.open_workbook(str(path))
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        for sheet_idx in range(xls.nsheets):
            xls_sheet = xls.sheet_by_index(sheet_idx)
            ws = wb.create_sheet(title=xls_sheet.name)
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    cell = xls_sheet.cell(row, col)
                    value = cell.value
                    # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error, 6=blank
                    if cell.ctype == 3:  # date
                        try:
                            from datetime import datetime as _dt
                            date_tuple = xlrd.xldate_as_tuple(value, xls.datemode)
                            value = _dt(*date_tuple)
                        except Exception:
                            pass
                    elif cell.ctype == 4:  # bool
                        value = bool(value)
                    elif cell.ctype in (0, 6):  # empty/blank
                        value = None
                    ws.cell(row=row + 1, column=col + 1, value=value)
        return wb

    elif suffix == ".csv":
        import csv as csv_mod
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = path.stem
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv_mod.reader(f)
            for row_idx, row_data in enumerate(reader, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    # Auto-detect types
                    v = value.strip()
                    if v == "":
                        continue
                    # Try int
                    try:
                        ws.cell(row=row_idx, column=col_idx, value=int(v))
                        continue
                    except ValueError:
                        pass
                    # Try float
                    try:
                        ws.cell(row=row_idx, column=col_idx, value=float(v))
                        continue
                    except ValueError:
                        pass
                    # String (including formulas starting with =)
                    ws.cell(row=row_idx, column=col_idx, value=v)
        return wb

    else:
        # .xlsx (default)
        return openpyxl.load_workbook(path, data_only=False)


def encrypt_excel(
    input_path: str,
    output_path: Optional[str] = None,
    rules: Optional[Dict[str, Any]] = None,
    scramble_numbers: bool = False,
    scramble_dates: bool = False,
) -> str:
    """
    Encrypt a spreadsheet file (.xlsx, .xls, or .csv).

    Default behavior:
    - ALL text → consistent substitution (same word = same fake word everywhere)
    - Numbers → kept unchanged (set scramble_numbers=True to scale them)
    - Dates → kept unchanged (set scramble_dates=True to offset them)
    - Formulas → preserved, text literals inside formulas also substituted

    Supports .xlsx, .xls (via xlrd), and .csv files.

    Returns:
        Path to the generated rules YAML file.
    """
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    default_encrypted, default_rules = generate_output_paths(input_path)
    if output_path is None:
        output_path = default_encrypted
    rules_path = str(
        Path(output_path).parent /
        (Path(output_path).stem.replace("_encrypted", "") + "_rules.yaml")
    )

    # ── 1. Open workbook (supports .xlsx, .xls, .csv) ─────────────────────
    wb = _load_workbook_any_format(src)

    # ── 2. First pass: collect ALL unique text values for global map ─────────
    all_texts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or cell.value == "":
                    continue
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        # Extract text literals from formulas
                        for literal in _extract_formula_literals(cell.value):
                            if literal and literal not in all_texts:
                                all_texts.append(literal)
                    else:
                        if cell.value not in all_texts:
                            all_texts.append(cell.value)

    global_text_map = build_global_text_map(all_texts)

    column_configs: Dict[str, Dict[str, Any]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        # ── 3. Build per-column transform configs ────────────────────────────
        sheet_configs: Dict[int, Dict[str, Any]] = {}

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            col_key = f"{sheet_name}!{col_letter}"

            # Collect ALL values in this column (all rows)
            col_values = []
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    col_values.append(cell.value)

            if rules and col_key in rules:
                cfg = dict(rules[col_key])
            else:
                transform = auto_detect_transform(col_values, scramble_numbers, scramble_dates)
                cfg = {"transform": transform}

                if transform == "scale":
                    cfg["factor"] = _random_factor()
                elif transform == "offset":
                    cfg["offset_days"] = _random_day_offset()
                    cfg["offset"] = _random_offset()

            column_configs[col_key] = cfg
            sheet_configs[col_idx] = cfg

        # ── 4. Extract formulas ──────────────────────────────────────────────
        formula_map = extract_formulas(ws)
        clear_formula_cells(ws, formula_map)

        # ── 5. Transform ALL cells in ALL rows ───────────────────────────────
        for row in range(1, max_row + 1):
            for col_idx, cfg in sheet_configs.items():
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is None:
                    continue
                try:
                    cell.value = _transform_cell(cell.value, cfg, global_text_map)
                except Exception:
                    pass

        # ── 6. Reinject formulas with text substitutions ─────────────────────
        for (row, col), formula in formula_map.items():
            updated = substitute_text_in_formula(formula, global_text_map)
            ws.cell(row=row, column=col).value = updated

    # ── 7. Save encrypted workbook ───────────────────────────────────────────
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # ── 8. Store global inverse text map in rules ────────────────────────────
    inverse_map = {v: k for k, v in global_text_map.items()}
    column_configs["__global_text_map"] = {
        "transform": "substitute",
        "mapping": inverse_map,
    }

    # ── 9. Generate and save rules ───────────────────────────────────────────
    metadata = {
        "original_filename": src.name,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "sheets": wb.sheetnames,
    }
    rules_dict = generate_rules(metadata, column_configs)
    save_rules(rules_dict, rules_path)

    return rules_path


def _extract_formula_literals(formula: str) -> list:
    """Extract all string literals (text between double quotes) from a formula."""
    literals = []
    i = 0
    chars = list(formula)
    while i < len(chars):
        if chars[i] == '"':
            i += 1
            start = i
            while i < len(chars) and chars[i] != '"':
                i += 1
            literal = "".join(chars[start:i])
            if literal:
                literals.append(literal)
            i += 1
        else:
            i += 1
    return literals
