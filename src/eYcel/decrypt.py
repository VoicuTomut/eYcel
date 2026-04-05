"""
Excel decryption pipeline.

Workflow
--------
1. Load and validate rules.yaml.
2. Open the encrypted workbook (data_only=False).
3. Extract the global inverse text map (fake→original).
4. For each sheet, extract formulas, reverse-transform ALL cells,
   reinject formulas with text literals reversed.
5. Save the restored workbook.
"""
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter


from .formula_handler import extract_formulas, clear_formula_cells
from .transformations import (
    reverse_offset_date,
    reverse_offset_number,
    reverse_scale,
    reverse_shuffle,
    transform_keep,
    reverse_text_in_formula,
)
from .yaml_handler import load_rules


# ---------------------------------------------------------------------------
# Reverse transform dispatcher
# ---------------------------------------------------------------------------

def _reverse_cell(value: Any, config: Dict[str, Any], inverse_text_map: Dict[str, str]) -> Any:
    """Apply the correct reverse transform to a single cell value."""
    # Text cells that exist in the global inverse map are always reversed,
    # regardless of column transform. This restores headers, titles, etc.
    if isinstance(value, str) and not value.startswith("="):
        if value in inverse_text_map:
            return inverse_text_map[value]

    t = config.get("transform", "keep")

    if t == "hash":
        return value  # one-way

    if t == "offset":
        if hasattr(value, "year"):
            return reverse_offset_date(value, config.get("offset_days", 0))
        return reverse_offset_number(float(value), config.get("offset", 0.0))

    if t == "scale":
        return reverse_scale(float(value), config["factor"])

    if t == "shuffle":
        return reverse_shuffle(str(value), config.get("mapping", {}))

    if t in ("anonymize", "keep"):
        return transform_keep(value)

    return value


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_and_validate_rules(rules_path: str) -> Dict[str, Any]:
    """Load a rules YAML file and raise on invalid content."""
    return load_rules(rules_path)


def apply_reverse_transform(
    column_data: List[Any],
    transform_config: Dict[str, Any],
) -> List[Any]:
    """Apply a reverse transform to an entire column's data list."""
    inverse = {}  # no global map needed for this legacy API
    return [
        _reverse_cell(v, transform_config, inverse) if v is not None else v
        for v in column_data
    ]


def decrypt_excel(
    encrypted_path: str,
    rules_path: str,
    output_path: str,
) -> None:
    """
    Restore an encrypted Excel file to its original values.

    Handles the global text substitution map for consistent reversal.
    """
    # ── 1. Load & validate rules ─────────────────────────────────────────────
    rules = load_and_validate_rules(rules_path)
    column_configs: Dict[str, Dict[str, Any]] = rules.get("columns", {})

    # ── 2. Extract global inverse text map ───────────────────────────────────
    global_entry = column_configs.get("__global_text_map", {})
    inverse_text_map: Dict[str, str] = global_entry.get("mapping", {})

    # ── 3. Open encrypted workbook ───────────────────────────────────────────
    src = Path(encrypted_path)
    if not src.exists():
        raise FileNotFoundError(f"Encrypted file not found: {encrypted_path}")

    wb = openpyxl.load_workbook(src, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        if max_row < 1:
            continue

        # ── 4. Extract formulas ──────────────────────────────────────────────
        formula_map = extract_formulas(ws)
        clear_formula_cells(ws, formula_map)

        # ── 5. Reverse-transform ALL cells ───────────────────────────────────
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            col_key = f"{sheet_name}!{col_letter}"

            cfg = column_configs.get(col_key)
            if cfg is None:
                continue
            if cfg.get("transform") == "substitute" and col_key == "__global_text_map":
                continue

            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is None:
                    continue
                try:
                    cell.value = _reverse_cell(cell.value, cfg, inverse_text_map)
                except Exception:
                    pass

        # ── 6. Reinject formulas with text reversed ──────────────────────────
        for (row, col), formula in formula_map.items():
            restored = reverse_text_in_formula(formula, inverse_text_map)
            ws.cell(row=row, column=col).value = restored

    # ── 7. Save restored workbook ────────────────────────────────────────────
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
