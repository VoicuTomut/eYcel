"""
Tests for column_analyzer module.
"""
import pytest
from datetime import date

import openpyxl

from src.eYcel.column_analyzer import (
    detect_cell_type,
    is_formula_cell,
    detect_categorical,
    get_column_stats,
    analyze_column,
    analyze_workbook_columns,
)


# ---------------------------------------------------------------------------
# detect_cell_type
# ---------------------------------------------------------------------------

class TestDetectCellType:
    def test_none_is_empty(self):
        assert detect_cell_type(None) == "empty"

    def test_empty_string_is_empty(self):
        assert detect_cell_type("") == "empty"

    def test_date(self):
        assert detect_cell_type(date(2023, 1, 1)) == "date"

    def test_int(self):
        assert detect_cell_type(42) == "int"

    def test_float(self):
        assert detect_cell_type(3.14) == "float"

    def test_string(self):
        assert detect_cell_type("hello") == "string"

    def test_bool(self):
        assert detect_cell_type(True) == "boolean"


# ---------------------------------------------------------------------------
# is_formula_cell
# ---------------------------------------------------------------------------

class TestIsFormulaCell:
    def _make_cell(self, value):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = value
        return ws["A1"]

    def test_formula_detected(self):
        cell = self._make_cell("=SUM(A1:A5)")
        assert is_formula_cell(cell) is True

    def test_plain_string_not_formula(self):
        cell = self._make_cell("hello")
        assert is_formula_cell(cell) is False

    def test_number_not_formula(self):
        cell = self._make_cell(123)
        assert is_formula_cell(cell) is False

    def test_none_not_formula(self):
        cell = self._make_cell(None)
        assert is_formula_cell(cell) is False


# ---------------------------------------------------------------------------
# detect_categorical
# ---------------------------------------------------------------------------

class TestDetectCategorical:
    def test_low_cardinality_is_categorical(self):
        values = ["A", "B", "A", "B", "A", "B", "A", "B", "C", "C"]
        assert detect_categorical(values) is True

    def test_high_cardinality_not_categorical(self):
        values = [f"item_{i}" for i in range(100)]
        assert detect_categorical(values) is False

    def test_empty_list(self):
        assert detect_categorical([]) is False

    def test_all_same_value(self):
        values = ["X"] * 20
        assert detect_categorical(values) is True


# ---------------------------------------------------------------------------
# get_column_stats
# ---------------------------------------------------------------------------

class TestGetColumnStats:
    def test_basic_stats(self):
        stats = get_column_stats([1.0, 2.0, 3.0, 4.0, 5.0])
        assert stats["min"] == 1.0
        assert stats["max"] == 5.0
        assert stats["avg"] == 3.0
        assert stats["unique_count"] == 5.0

    def test_empty_list(self):
        stats = get_column_stats([])
        assert stats["min"] == 0.0
        assert stats["max"] == 0.0

    def test_single_value(self):
        stats = get_column_stats([7.0])
        assert stats["min"] == 7.0
        assert stats["max"] == 7.0
        assert stats["avg"] == 7.0


# ---------------------------------------------------------------------------
# analyze_column  (requires a real worksheet)
# ---------------------------------------------------------------------------

class TestAnalyzeColumn:
    def _make_wb_with_col(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "amount"
        ws["A2"] = 100.0
        ws["A3"] = 200.0
        ws["A4"] = 150.0
        ws["A5"] = 300.0
        ws["A6"] = 250.0
        ws["A7"] = 175.0
        ws["B1"] = "category"
        # Repeat only 2 unique values across 6 rows → ratio = 2/6 ≈ 0.33 → categorical
        ws["B2"] = "Electronics"
        ws["B3"] = "Clothing"
        ws["B4"] = "Electronics"
        ws["B5"] = "Clothing"
        ws["B6"] = "Electronics"
        ws["B7"] = "Clothing"
        ws["C1"] = "total"
        ws["C2"] = "=A2*1.1"
        ws["C3"] = "=A3*1.1"
        return wb, ws

    def test_numeric_column(self):
        wb, ws = self._make_wb_with_col()
        meta = analyze_column(ws, "A")
        assert meta["dominant_type"] in ("float", "int")
        assert meta["data_count"] == 6   # 6 data rows added
        assert meta["formula_count"] == 0

    def test_formula_column(self):
        wb, ws = self._make_wb_with_col()
        meta = analyze_column(ws, "C")
        assert meta["formula_count"] == 2
        assert meta["data_count"] == 0

    def test_categorical_column(self):
        wb, ws = self._make_wb_with_col()
        meta = analyze_column(ws, "B")
        assert meta["dominant_type"] == "categorical"

    def test_header_captured(self):
        wb, ws = self._make_wb_with_col()
        meta = analyze_column(ws, "A")
        assert meta["header"] == "amount"


# ---------------------------------------------------------------------------
# analyze_workbook_columns
# ---------------------------------------------------------------------------

class TestAnalyzeWorkbookColumns:
    def test_returns_all_columns(self, sample_excel_path):
        wb = openpyxl.load_workbook(str(sample_excel_path), data_only=False)
        result = analyze_workbook_columns(wb)
        # sample has 5 columns A-E
        assert len(result) == 5

    def test_column_letters_as_keys(self, sample_excel_path):
        wb = openpyxl.load_workbook(str(sample_excel_path), data_only=False)
        result = analyze_workbook_columns(wb)
        assert "A" in result
        assert "E" in result
