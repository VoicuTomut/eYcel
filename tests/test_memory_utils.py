"""
Tests for memory_utils module.
"""
from __future__ import annotations

import pytest
import openpyxl

from src.eYcel.memory_utils import (
    get_memory_usage_mb,
    check_memory_limit,
    chunk_iterator,
    process_column_in_chunks,
    DEFAULT_CHUNK_SIZE,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_worksheet(n_rows: int, n_cols: int = 3) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return a worksheet with *n_rows* data rows (row 1 = headers)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # header row
    ws.append([f"col{c}" for c in range(1, n_cols + 1)])
    # data rows
    for i in range(1, n_rows + 1):
        ws.append([i * (c + 1) for c in range(n_cols)])
    return ws


# ---------------------------------------------------------------------------
# get_memory_usage_mb
# ---------------------------------------------------------------------------

class TestGetMemoryUsageMb:
    def test_returns_positive_float(self):
        mb = get_memory_usage_mb()
        # May be 0.0 on unsupported platforms but must not raise
        assert isinstance(mb, float)
        assert mb >= 0.0

    def test_returns_reasonable_value(self):
        mb = get_memory_usage_mb()
        # Python process should use at least 1 MB and less than 2 GB
        # (0.0 is also acceptable on unsupported platforms)
        assert mb < 2048


# ---------------------------------------------------------------------------
# check_memory_limit
# ---------------------------------------------------------------------------

class TestCheckMemoryLimit:
    def test_no_warning_under_limit(self, capsys):
        check_memory_limit(max_mb=99999.0, label="test")
        captured = capsys.readouterr()
        assert captured.err == ""

    def test_warning_when_over_limit(self, capsys):
        check_memory_limit(max_mb=0.0, label="mytest")
        captured = capsys.readouterr()
        assert "WARNING" in captured.err

    def test_label_included_in_warning(self, capsys):
        check_memory_limit(max_mb=0.0, label="phase5")
        captured = capsys.readouterr()
        assert "phase5" in captured.err


# ---------------------------------------------------------------------------
# chunk_iterator
# ---------------------------------------------------------------------------

class TestChunkIterator:
    def test_all_rows_covered(self):
        n_rows = 25
        ws = _make_worksheet(n_rows)
        seen = 0
        for chunk in chunk_iterator(ws, chunk_size=10):
            seen += len(chunk)
        assert seen == n_rows

    def test_chunk_size_respected(self):
        ws = _make_worksheet(50)
        sizes = [len(chunk) for chunk in chunk_iterator(ws, chunk_size=10)]
        # All but possibly the last chunk should have 10 rows
        for s in sizes[:-1]:
            assert s == 10

    def test_last_chunk_smaller(self):
        ws = _make_worksheet(23)
        chunks = list(chunk_iterator(ws, chunk_size=10))
        assert len(chunks[-1]) == 3

    def test_empty_sheet_yields_nothing(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["header"])   # only header row
        chunks = list(chunk_iterator(ws))
        assert chunks == []

    def test_single_row_yields_one_chunk(self):
        ws = _make_worksheet(1)
        chunks = list(chunk_iterator(ws))
        assert len(chunks) == 1
        assert len(chunks[0]) == 1

    def test_skips_header_by_default(self):
        ws = _make_worksheet(5)
        first_val = ws.cell(row=2, column=1).value  # first data cell
        chunks = list(chunk_iterator(ws))
        first_yielded_row = chunks[0][0]  # first row in first chunk
        assert first_yielded_row[0].value == first_val


# ---------------------------------------------------------------------------
# process_column_in_chunks
# ---------------------------------------------------------------------------

class TestProcessColumnInChunks:
    def test_transforms_all_cells(self):
        ws = _make_worksheet(20, n_cols=2)
        count = process_column_in_chunks(ws, column_index=1, processor_func=lambda v: v * 2)
        assert count == 20

    def test_values_actually_changed(self):
        ws = _make_worksheet(5, n_cols=1)
        original_vals = [ws.cell(row=r, column=1).value for r in range(2, 7)]
        process_column_in_chunks(ws, column_index=1, processor_func=lambda v: v + 100)
        new_vals = [ws.cell(row=r, column=1).value for r in range(2, 7)]
        for orig, new in zip(original_vals, new_vals):
            assert new == orig + 100

    def test_formula_cells_not_touched(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["value"])
        ws.append(["=1+1"])
        ws.append([42])
        process_column_in_chunks(ws, column_index=1, processor_func=lambda v: v * 2)
        assert ws.cell(row=2, column=1).value == "=1+1"  # unchanged
        assert ws.cell(row=3, column=1).value == 84      # doubled

    def test_none_cells_not_touched(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["value"])
        ws.append([None])
        ws.append([10])
        count = process_column_in_chunks(ws, column_index=1, processor_func=lambda v: v + 1)
        assert count == 1  # only the non-None cell
        assert ws.cell(row=2, column=1).value is None

    def test_returns_processed_count(self):
        ws = _make_worksheet(7)
        count = process_column_in_chunks(ws, column_index=1, processor_func=lambda v: v)
        assert count == 7

    def test_chunked_gives_same_result_as_whole(self):
        ws_whole = _make_worksheet(30)
        ws_chunk = _make_worksheet(30)
        process_column_in_chunks(ws_whole, 1, lambda v: v * 3, chunk_size=30)
        process_column_in_chunks(ws_chunk, 1, lambda v: v * 3, chunk_size=5)
        for r in range(2, 32):
            assert ws_whole.cell(row=r, column=1).value == ws_chunk.cell(row=r, column=1).value
