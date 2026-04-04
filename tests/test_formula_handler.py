"""
Tests for formula_handler module.
"""
import pytest
import openpyxl

from src.eYcel.formula_handler import (
    extract_formulas,
    clear_formula_cells,
    reinject_formulas,
    verify_formulas_preserved,
    get_formula_summary,
)


def _make_ws_with_formulas():
    """Return a workbook+worksheet containing data and formula cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "value"
    ws["B1"] = "double"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["B2"] = "=A2*2"
    ws["B3"] = "=A3*2"
    ws["B4"] = "=SUM(B2:B3)"
    return wb, ws


class TestExtractFormulas:
    def test_finds_all_formulas(self):
        _, ws = _make_ws_with_formulas()
        formulas = extract_formulas(ws)
        assert len(formulas) == 3

    def test_keys_are_row_col_tuples(self):
        _, ws = _make_ws_with_formulas()
        formulas = extract_formulas(ws)
        for key in formulas:
            assert isinstance(key, tuple)
            assert len(key) == 2

    def test_values_start_with_equals(self):
        _, ws = _make_ws_with_formulas()
        formulas = extract_formulas(ws)
        for formula in formulas.values():
            assert formula.startswith("=")

    def test_data_cells_not_captured(self):
        _, ws = _make_ws_with_formulas()
        formulas = extract_formulas(ws)
        # A2 and A3 are data, should not appear
        assert (2, 1) not in formulas
        assert (3, 1) not in formulas

    def test_empty_sheet_returns_empty(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert extract_formulas(ws) == {}


class TestClearFormulasCells:
    def test_formulas_blanked(self):
        _, ws = _make_ws_with_formulas()
        formulas = extract_formulas(ws)
        clear_formula_cells(ws, formulas)
        for (row, col) in formulas:
            assert ws.cell(row=row, column=col).value is None

    def test_data_cells_untouched(self):
        _, ws = _make_ws_with_formulas()
        formulas = extract_formulas(ws)
        clear_formula_cells(ws, formulas)
        assert ws["A2"].value == 10
        assert ws["A3"].value == 20


class TestReinjectFormulas:
    def test_formulas_restored(self):
        _, ws = _make_ws_with_formulas()
        formulas = extract_formulas(ws)
        clear_formula_cells(ws, formulas)
        reinject_formulas(ws, formulas)
        for (row, col), formula in formulas.items():
            assert ws.cell(row=row, column=col).value == formula


class TestVerifyFormulasPreserved:
    def test_identical_sheets_pass(self):
        wb1, ws1 = _make_ws_with_formulas()
        wb2, ws2 = _make_ws_with_formulas()
        assert verify_formulas_preserved(ws1, ws2) is True

    def test_missing_formula_fails(self):
        wb1, ws1 = _make_ws_with_formulas()
        wb2, ws2 = _make_ws_with_formulas()
        ws2["B2"] = None        # remove one formula
        assert verify_formulas_preserved(ws1, ws2) is False

    def test_changed_formula_fails(self):
        wb1, ws1 = _make_ws_with_formulas()
        wb2, ws2 = _make_ws_with_formulas()
        ws2["B2"] = "=A2*99"   # different formula
        assert verify_formulas_preserved(ws1, ws2) is False


class TestGetFormulaSummary:
    def test_counts_correct(self):
        _, ws = _make_ws_with_formulas()
        summary = get_formula_summary(ws)
        assert summary["formula_count"] == 3
        # headers + data cells: A1, B1, A2, A3
        assert summary["data_count"] == 4
