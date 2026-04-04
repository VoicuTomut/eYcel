"""
Shared pytest fixtures for the eYcel test suite.
"""
import pytest
import tempfile
from pathlib import Path
from datetime import date

import openpyxl


@pytest.fixture
def temp_dir():
    """Provide a temporary directory that is cleaned up after the test."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_excel_path(temp_dir):
    """
    Create a minimal sample Excel file for testing.

    Sheet layout (row 1 = headers):
      A: customer_id (string IDs)
      B: amount      (floats)
      C: category    (categorical strings)
      D: joined_date (dates)
      E: total       (formula =B2*1.1 etc.)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Headers
    ws["A1"] = "customer_id"
    ws["B1"] = "amount"
    ws["C1"] = "category"
    ws["D1"] = "joined_date"
    ws["E1"] = "total"

    # Data rows
    rows = [
        ("CUST001", 100.0,  "Electronics", date(2022, 1, 15), "=B2*1.1"),
        ("CUST002", 250.0,  "Clothing",    date(2022, 3, 10), "=B3*1.1"),
        ("CUST003", 75.50,  "Electronics", date(2021, 7, 22), "=B4*1.1"),
        ("CUST004", 320.00, "Food",        date(2023, 5, 5),  "=B5*1.1"),
        ("CUST005", 50.0,   "Clothing",    date(2020, 11, 1), "=B6*1.1"),
    ]
    for row in rows:
        ws.append(row)

    path = temp_dir / "sample.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_rules():
    """Return a minimal valid rules dictionary."""
    return {
        "metadata": {
            "original_filename": "test.xlsx",
            "timestamp": "2024-01-15T10:00:00+00:00",
            "sheets": ["Sheet1"],
        },
        "columns": {
            "customer_id": {"transform": "hash", "salt": "testsalt123"},
            "amount":       {"transform": "scale", "factor": 0.5},
            "category":     {"transform": "shuffle", "mapping": {"Electronics": "Cat_0", "Clothing": "Cat_1", "Food": "Cat_2"}},
            "joined_date":  {"transform": "offset", "offset_days": -30, "offset": 0.0},
        },
    }
