"""
Tests for the encrypt module.
"""
import pytest
from pathlib import Path
import openpyxl

from src.eYcel.encrypt import (
    encrypt_excel,
    auto_detect_transform,
    generate_output_paths,
)


class TestGenerateOutputPaths:
    def test_encrypted_suffix(self):
        enc, _ = generate_output_paths("/tmp/sales.xlsx")
        assert enc.endswith("_encrypted.xlsx")

    def test_rules_suffix(self):
        _, rules = generate_output_paths("/tmp/sales.xlsx")
        assert rules.endswith("_rules.yaml")

    def test_same_directory(self):
        enc, rules = generate_output_paths("/data/myfile.xlsx")
        assert Path(enc).parent == Path(rules).parent


class TestAutoDetectTransform:
    def test_date_column_default_keeps(self):
        from datetime import date
        values = [date(2024, 1, 1), date(2024, 6, 15)]
        assert auto_detect_transform(values) == "keep"

    def test_date_column_scramble(self):
        from datetime import date
        values = [date(2024, 1, 1), date(2024, 6, 15)]
        assert auto_detect_transform(values, scramble_dates=True) == "offset"

    def test_float_column_default_keeps(self):
        values = [1.5, 2.7, 3.9]
        assert auto_detect_transform(values) == "keep"

    def test_float_column_scramble(self):
        values = [1.5, 2.7, 3.9]
        assert auto_detect_transform(values, scramble_numbers=True) == "scale"

    def test_int_column_default_keeps(self):
        values = [1, 2, 3]
        assert auto_detect_transform(values) == "keep"

    def test_string_column(self):
        values = ["Alice", "Bob", "Carol"]
        assert auto_detect_transform(values) == "substitute"

    def test_empty_keeps(self):
        values = [None, "", None]
        assert auto_detect_transform(values) == "keep"


class TestEncryptExcel:
    def test_creates_output_file(self, sample_excel_path, temp_dir):
        out = str(temp_dir / "enc.xlsx")
        encrypt_excel(str(sample_excel_path), out)
        assert Path(out).exists()

    def test_creates_rules_file(self, sample_excel_path, temp_dir):
        out = str(temp_dir / "enc.xlsx")
        rules_path = encrypt_excel(str(sample_excel_path), out)
        assert Path(rules_path).exists()

    def test_formulas_preserved(self, sample_excel_path, temp_dir):
        out = str(temp_dir / "enc.xlsx")
        encrypt_excel(str(sample_excel_path), out)
        wb_orig = openpyxl.load_workbook(str(sample_excel_path), data_only=False)
        wb_enc = openpyxl.load_workbook(out, data_only=False)
        # Check formula cells in column E
        for row in range(2, 7):
            orig_val = wb_orig.active.cell(row=row, column=5).value
            enc_val = wb_enc.active.cell(row=row, column=5).value
            if orig_val and str(orig_val).startswith("="):
                assert enc_val == orig_val, f"Formula at row {row} changed!"

    def test_text_values_changed(self, sample_excel_path, temp_dir):
        out = str(temp_dir / "enc.xlsx")
        encrypt_excel(str(sample_excel_path), out)
        wb_orig = openpyxl.load_workbook(str(sample_excel_path), data_only=False)
        wb_enc = openpyxl.load_workbook(out, data_only=False)
        # Text values in col A (names) should be substituted
        orig_vals = [wb_orig.active.cell(row=r, column=1).value for r in range(2, 7)]
        enc_vals  = [wb_enc.active.cell(row=r, column=1).value  for r in range(2, 7)]
        assert orig_vals != enc_vals, "Text values should be substituted"

    def test_numbers_kept_by_default(self, sample_excel_path, temp_dir):
        out = str(temp_dir / "enc.xlsx")
        encrypt_excel(str(sample_excel_path), out)
        wb_orig = openpyxl.load_workbook(str(sample_excel_path), data_only=False)
        wb_enc = openpyxl.load_workbook(out, data_only=False)
        # Numbers in col B (amount) should be kept unchanged by default
        orig_vals = [wb_orig.active.cell(row=r, column=2).value for r in range(2, 7)]
        enc_vals  = [wb_enc.active.cell(row=r, column=2).value  for r in range(2, 7)]
        assert orig_vals == enc_vals, "Numbers should be kept unchanged by default"

    def test_missing_input_raises(self, temp_dir):
        with pytest.raises(FileNotFoundError):
            encrypt_excel(str(temp_dir / "no_such_file.xlsx"), str(temp_dir / "out.xlsx"))

    def test_headers_substituted(self, sample_excel_path, temp_dir):
        """Headers are now substituted too (hides column meaning)."""
        out = str(temp_dir / "enc.xlsx")
        encrypt_excel(str(sample_excel_path), out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Headers should be substituted (NOT the original text)
        assert ws["A1"].value != "customer_id", "Header should be substituted"
        assert ws["B1"].value != "amount", "Header should be substituted"
        # But they should be non-empty strings
        assert isinstance(ws["A1"].value, str) and len(ws["A1"].value) > 0
